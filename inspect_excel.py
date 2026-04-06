import sys
from pathlib import Path
from typing import List, Literal

import pandas as pd
import openpyxl


def print_header(title: str) -> None:
    print(title)
    print('-' * len(title))


def inspect_excel_file(file_path: Path) -> None:
    if not file_path.exists():
        print(f'Error: file not found: {file_path}')
        return

    print_header('File details')
    print(f'Path: {file_path.resolve()}')
    print(f'Exists: {file_path.exists()}')
    print(f'Size: {file_path.stat().st_size} bytes')
    print()

    print_header('Trying pandas read options')
    extension = file_path.suffix.lower()
    engines: List[Literal['openpyxl', 'xlrd']] = []
    if extension == '.xlsx':
        engines = ['openpyxl']
    elif extension == '.xls':
        engines = ['xlrd']
    else:
        engines = ['openpyxl', 'xlrd']

    for engine in engines:
        try:
            df = pd.read_excel(file_path, engine=engine)
            print(f'{engine} engine shape: {df.shape}')
        except Exception as e:
            print(f'{engine} error: {e}')

    print()
    try:
        xls = pd.ExcelFile(file_path)
        print(f'Sheets in file: {xls.sheet_names}')

        for sheet in xls.sheet_names:
            try:
                df = pd.read_excel(xls, sheet_name=sheet)
                print(f'Sheet "{sheet}" shape: {df.shape}')
                print(f'Columns: {list(df.columns)}')
                print('First 5 rows:')
                print(df.head().to_string(index=False))
                print()
            except Exception as e:
                print(f'Error reading sheet "{sheet}": {e}')
    except Exception as e:
        print(f'Could not inspect sheets: {e}')

    print_header('Checking raw data')
    try:
        df_raw = pd.read_excel(file_path, header=None)
        print(f'Raw shape: {df_raw.shape}')
        for row_num, (_idx, row) in enumerate(df_raw.iterrows(), start=1):
            print(f'Row {row_num}: {list(row.values)}')
    except Exception as e:
        print(f'Error reading raw data: {e}')

    print()
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        if ws is None:
            print('Openpyxl workbook has no active worksheet')
            return

        print_header('Openpyxl detailed inspection')
        print(f'Worksheet title: {ws.title}')
        print(f'Max row: {ws.max_row}, Max col: {ws.max_column}')
        print()
        print('All cell values:')
        for row in ws.iter_rows(values_only=True):
            print(list(row))
    except Exception as e:
        print(f'Openpyxl workbook error: {e}')


if __name__ == '__main__':
    filename = sys.argv[1] if len(sys.argv) > 1 else 'BMS_BOM_Base.xlsx'
    inspect_excel_file(Path(filename))